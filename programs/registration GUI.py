import openpyxl
from tkinter import*
from openpyxl import*
wb=load_workbook('C:\\Users\\ADMIN\\Desktop\\excel.xlsx')
s=wb.active

def excel():
    s.column_dimensions['A'].width = 25
    s.column_dimensions['B'].width = 10
    s.column_dimensions['C'].width = 10
    s.column_dimensions['D'].width = 20
    s.column_dimensions['E'].width = 20
    s.column_dimensions['F'].width = 40
    s.column_dimensions['G'].width = 60

    s.cell(row=1, column=1).value = "Name"
    s.cell(row=1, column=2).value = "Course"
    s.cell(row=1, column=3).value = "Semister"
    s.cell(row=1, column=4).value = "Form Number"
    s.cell(row=1, column=5).value = "Contact Number"
    s.cell(row=1, column=6).value = "Email ID"
    s.cell(row=1, column=7).value = "Address"

def f1(event):
    co.focus_set()

def f2(event):
    sem.focus_set()

def f3(event):
    fno.focus_set()

def f4(event):
    cno.focus_set()

def f5(event):
    eid.focus_set()

def f6(event):
    add.focus_set()

def clear():
    na.delete(0, END)
    co.delete(0, END)
    sem.delete(0, END)
    fno.delete(0, END)
    cno.delete(0, END)
    eid.delete(0, END)
    add.delete(0, END)

def insert():
    if(na.get()=="" and co.get()=="" and fno.get()=="" and sem.get()=="" and cno.get()=="" and eid.get()=="" and add.get()==""):
        print("Empty input!!!")
    else:
        cr=s.max_row
        cc=s.max_column

        s.cell(row=cr + 1, column=1).value = na.get()
        s.cell(row=cr + 1, column=2).value = co.get()
        s.cell(row=cr + 1, column=3).value = sem.get()
        s.cell(row=cr + 1, column=4).value = fno.get()
        s.cell(row=cr + 1, column=5).value = cno.get()
        s.cell(row=cr + 1, column=6).value = eid.get()
        s.cell(row=cr + 1, column=7).value = add.get()

        wb.save('C:\\Users\\ADMIN\\Desktop\\excel.xlsx')
        na.focus_set()
        clear()

if __name__=="__main__":
    root=Tk()
    root.configure(background="gold")
    root.title("Registration Form")
    root.geometry("500x300")
    excel()
    heading = Label(root, text="FORM", bg="gold")
    heading.config(font=("Algerian", 25, "bold"))
    na = Label(root, text="Name", bg="gold")
    na.config(font=("Copperplate Gothic Bold", 10))
    co = Label(root, text="Course", bg="gold")
    co.config(font=("Copperplate Gothic Bold", 10))
    sem = Label(root, text="Semester", bg="gold")
    sem.config(font=("Copperplate Gothic Bold", 10))
    fno = Label(root, text="Form Number", bg="gold")
    fno.config(font=("Copperplate Gothic Bold", 10))
    cno = Label(root, text="Contact Number", bg="gold")
    cno.config(font=("Copperplate Gothic Bold", 10))
    eid = Label(root, text="Email ID", bg="gold")
    eid.config(font=("Copperplate Gothic Bold", 10))
    add = Label(root, text="Address", bg="gold")
    add.config(font=("Copperplate Gothic Bold", 10))

    heading.grid(row=0, column=1)
    na.grid(row=1, column=0)
    co.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    fno.grid(row=4, column=0)
    cno.grid(row=5, column=0)
    eid.grid(row=6, column=0)
    add.grid(row=7, column=0)

    na = Entry(root)
    co = Entry(root)
    sem = Entry(root)
    fno = Entry(root)
    cno = Entry(root)
    eid = Entry(root)
    add = Entry(root)

    na.bind("<Return>", f1)
    co.bind("<Return>", f2)
    sem.bind("<Return>", f3)
    fno.bind("<Return>", f4)
    cno.bind("<Return>", f5)
    eid.bind("<Return>", f6)

    na.grid(row=1, column=1, ipadx="100")
    co.grid(row=2, column=1, ipadx="100")
    sem.grid(row=3, column=1, ipadx="100")
    fno.grid(row=4, column=1, ipadx="100")
    cno.grid(row=5, column=1, ipadx="100")
    eid.grid(row=6, column=1, ipadx="100")
    add.grid(row=7, column=1, ipadx="100")

    excel()
    submit=Button(root,text="Submit",fg="Black",bg="Red",command=insert)
    submit.grid(row=8,column=1)
    root.mainloop()