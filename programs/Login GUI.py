from openpyxl import*
from tkinter import*
w=load_workbook("C:\\Users\\Sivant\\Desktop\\pass.xlsx")
p=w.active

def password():
    p.column_dimensions['A'].width = 25
    p.column_dimensions['B'].width = 40
    p.cell(row=1,column=1).value="Username"
    p.cell(row=1,column=2).value="Password"

def f1(event):
    pa.focus_set()

def clear():
    un.delete(0,END)
    pa.delete(0,END)

def insert():
    if(pa.get()=="" and un.get()==""):
        print("empty input!!!")
    else:
        cr=p.max_row
        cp=p.max_column
        p.cell(row=cr + 1, column=1).value = un.get()
        p.cell(row=cr + 1, column=2).value = pa.get()
        w.save("C:\\Users\\Sivant\\Desktop\\pass.xlsx")
        un.focus_set()
        clear()

if __name__=="__main__":
    r=Tk()
    r.configure(background="Pink")
    r.geometry("400x200")
    r.title("Login")
    heading= Label(r,text="Login",bg="pink")
    heading.configure(font=("Algerian",30,"bold"))
    un = Label(r, text="Username", bg="pink")
    un.configure(font=("Copperplate Gothic Bold", 10),fg="blue")
    pa = Label(r, text="Password", bg="pink")
    pa.configure(font=("Copperplate Gothic Bold", 10),fg="blue")

    heading.grid(row=0, column=1)
    un.grid(row=1, column=0)
    pa.grid(row=2, column=0)

    un = Entry(r)
    pa = Entry(r)

    pa.bind("<Return>",f1)

    un.grid(row=1, column=1, ipadx="80")
    pa.grid(row=2, column=1, ipadx="80")

    password()
    space = Label(r, text="", bg="pink")
    space.grid(row=3,column=1)
    submit=Button(r,text="Submit",fg="black",command=insert)
    submit.grid(row=4,column=1)
    r.mainloop()
    